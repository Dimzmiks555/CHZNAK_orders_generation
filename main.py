import sys
import openpyxl
import re
import xml.etree.ElementTree as ET
import copy
import uuid

def main():
    processname = sys.argv[0]
    filename = sys.argv[1]
    wookbook = openpyxl.load_workbook('import/' + filename)
    worksheet = wookbook.active

    data = []

    # Iterate the loop to read the cell values
    for i in range(0, worksheet.max_row):

        row = []

        if i != 0:
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if col[0].value == 'Код с текстом':
                    code = re.findall("\\d+", col[i].value)[0]
                    row.append('0' + code)
                if col[0].value == 'Количество':
                    row.append(col[i].value)
            data.append(row)

    # print(data)

    xmldata = """<?xml version="1.0" encoding="utf-8"?>
<order xmlns="urn:oms.order" xsi:schemaLocation="urn:oms.order schema.xsd" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <lp>
    <productGroup>lp</productGroup>
    <contactPerson>ФИО</contactPerson>
    <releaseMethodType>REMAINS</releaseMethodType>
    <createMethodType>SELF_MADE</createMethodType>
    <productionOrderId></productionOrderId>
    <products>
      <product>		
        <gtin></gtin>				
        <quantity></quantity>		
        <serialNumberType>OPERATOR</serialNumberType>
		    <cisType>UNIT</cisType>
        <templateId>10</templateId>
      </product>
    </products>
  </lp>
</order>"""

    template = """<?xml version="1.0" encoding="utf-8"?>
    <order xmlns="urn:oms.order" xsi:schemaLocation="urn:oms.order schema.xsd" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <lp>
        <productGroup>lp</productGroup>
        <contactPerson>ФИО</contactPerson>
        <releaseMethodType>REMAINS</releaseMethodType>
        <createMethodType>SELF_MADE</createMethodType>
        <productionOrderId></productionOrderId>
        <products></products>
    </lp>
    </order>"""


    ET.register_namespace("","urn:oms.order")

    tree = ET.ElementTree(ET.fromstring(xmldata))
    root = tree.getroot()


    xmlparts = []

    for row in data:
        product = copy.deepcopy(root[0][5][0])
        product[0].text = row[0]
        product[1].text = str(row[1])
        xmlparts.append(product) 



    productCounter = 0

    files = []

    for product in xmlparts:
        if productCounter == 0:
            tree = ET.ElementTree(ET.fromstring(template))
            root = tree.getroot()
            root[0][4].text = str(uuid.uuid1())
            root[0][5].append(product)
            files.append(tree)
            productCounter = productCounter + 1
        elif productCounter < 9:
            root = files[len(files) - 1].getroot()
            root[0][5].append(product)
            productCounter = productCounter + 1
        elif productCounter == 9:
            root = files[len(files) - 1].getroot()
            root[0][5].append(product)
            productCounter = 0

    print(files)

    for index, file in enumerate(files):
        file.write(f'export/output_{index}.xml', encoding='utf-8')

if __name__ == '__main__':
    main() 
